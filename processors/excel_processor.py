import openpyxl
from openpyxl import load_workbook
import logging

from processors.type_assessor import detect_type

# Configuration des logs
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

logger = logging.getLogger(__name__)


def open_xlsx_file(path: str) -> dict:
    logger.info("Ouverture du fichier xlsx : %s", path)

    columns = {}

    wb = openpyxl.load_workbook(path, data_only=True)

    ws = wb.active

    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            debut_ligne = row[0].row
            print(f"Début du tableau : ligne {debut_ligne}")
            break

    return columns


def get_columns_types(columns: dict) -> list:
    logger.info("Début de la détection des types")

    types = []

    for col_name in columns:
        logger.info(
            "Analyse de la colonne '%s' (%s valeurs)",
            col_name,
            len(columns[col_name])
        )

        for value in columns[col_name]:
            detected_type = detect_type(value)
            types.append(f"{col_name}&&{detected_type}")

        logger.debug(
            "Types détectés pour '%s' : %s",
            col_name,
            set(
                detect_type(value)
                for value in columns[col_name]
            )
        )

    logger.info(
        "Détection terminée : %s types analysés",
        len(types)
    )

    return types